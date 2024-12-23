# -*- coding: utf-8 -*-
"""
Created on Fri Apr 26 18:57:13 2024

@author: admin
"""

# Define a function that accepts roll number and returns whether the student is present or absent.
# import random
# def attendence():
#     name = input("Enter the name : ")
#     roll_number = int(input("Enter the name : "))
#     n = random.randint(0, 1)
#     if n == 0:
#         print("Student is Present ")
#     else:
#         print("Student is Not Present")
# attendence()


# def area(radius):
#     return 3.14 * radius * radius

# radius = int(input("Enter the radius : "))
# print("area is : ", area(radius))
    
# Write a Python program to reverse a number

# n = 1234
# rev = 0
# while n > 0:
#     rem = n % 10
#     rev = rev * 10 + rem
#     n = n//10
# print(rev)



# Write a function called favorite_book() that accepts two parameter, title and author. The function should print a message, such as The History of Bhutan by Karma Phuntsho. Call the function, make sure to include a book title and author as an argument in the function call.
# def favorite_book(title,author):
#     print(f"{title} by {author}")
# favorite_book("Plassey To Partition", "Shekhar Bandhopadhyay")


# def leap_year(year):
#     if year % 4 == 0 or year % 400 == 0:
#         print(f"{year} is a leap year")
#     else:
#         print(f"{year} is not a leap year")


# def largest(a,b,c):
#     if a > c and a>b:
#         print(f"{a} is largest")
#     elif b > c and b > a:
#         print(f"{b} is largest")
#     else:
        
#         print(f"{c} is largest")
# largest(10, 20, 5)
  
# Write a function, is_vowel that returns the value true if a given character is a vowel, and otherwise returns false. Write another function main, in main() function accept a string from user and count number of vowels in that string.      
# vowel = "aeiouAEIOU"
# def is_vowel(ch):
#     if ch in vowel:
#         return True
#     else:
#         return False
# ch = input("enter a character : ")
# if is_vowel(ch):
#     print("character")
# else:
#     print("No Character")
       
        
    
# def vowel_count(str1):
#     count = 0
#     for i in str1:
#         if is_vowel(i):
#             count += 1
#     print("count is ", count)
    
# str1 = input("Enter the String : ")
# vowel_count(str1)

# place = ["Satara","Pimpri-Chinchwad","Kolhapur","Pune","Mumbai"]
# res = []

# def methood_count(place):
    

#     for i in place:
#         if len(i)>5:
#             res.append(i)
#     print(res)
# methood_count(place)


# from openpyxl import Workbook  
# wb = Workbook()  
# sheet = wb.active  
  
# rows_count = (  
#     (14, 27),  
#     (22, 30),  
#     (42, 92),  
#     (51, 32),  
#     (16, 60),  
#     (63, 13)  
# )  
  
# for i in rows_count:  
#     sheet.append(i)  
  
# cell = sheet.cell(row=7, column=3)  
# cell.value = "=SUM(A1:B6)"  
# cell.font = cell.font.copy(bold=True)  
# # cell.font = copy(cell.font, bold=True)
  
# wb.save('formulas_book.xlsx')  

from openpyxl import Workbook  
from openpyxl.styles import Font

wb = Workbook()  
sheet = wb.active  
  
rows_count = (  
    (14, 27),  
    (22, 30),  
    (42, 92),  
    (51, 32),  
    (16, 60),  
    (63, 13)  
)  
  
for i in rows_count:  
    sheet.append(i)  
  
cell = sheet.cell(row=7, column=3)  
cell.value = "=SUM(A1:B6)"  
cell.font = Font(bold=True)  # Using Font directly to set bold
  
wb.save('formulas_book.xlsx')
 